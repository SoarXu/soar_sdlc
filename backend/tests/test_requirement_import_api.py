from io import BytesIO
from uuid import uuid4

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.core.security import get_password_hash
from app.db.session import SessionLocal
from app.models.user import User
from app.models.iteration import Iteration
from app.models.work_item_iteration_history import WorkItemIterationHistory
from app.models.workflow_definition import WorkflowState
from app.services.requirement_import_service import REQUIREMENT_IMPORT_COLUMNS


def _xlsx(rows: list[list[str | None]]) -> tuple[str, bytes, str]:
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return (
        "requirements.xlsx",
        buffer.getvalue(),
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def test_requirement_import_template_downloads_excel(client: TestClient):
    response = client.get("/api/v1/requirements/import/template")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    assert [cell.value for cell in sheet[1]] == [
        "项目名称",
        "需求标题",
        "类型",
        "优先级",
        "提出人",
        "需求描述",
        "验收标准",
    ]
    assert sheet["A2"].value == "示例项目"
    assert sheet["C2"].value == "功能"
    assert sheet["D2"].value == "3"
    validations = list(sheet.data_validations.dataValidation)
    assert any("功能,接口,性能,安全,体验,改进,其他" in item.formula1 for item in validations)
    assert any("1,2,3,4,5" in item.formula1 for item in validations)


def test_requirement_import_template_prefills_project_name(client: TestClient):
    project_name = f"模板项目-{uuid4().hex[:8]}"
    project = client.post("/api/v1/projects", json={"name": project_name}).json()

    response = client.get(f"/api/v1/requirements/import/template?project_id={project['id']}")

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook.active
    assert sheet["A2"].value == project_name


def test_requirement_import_preview_rejects_non_excel(client: TestClient):
    response = client.post(
        "/api/v1/requirements/import/preview",
        files={"file": ("requirements.csv", b"title", "text/csv")},
    )

    assert response.status_code == 422


def test_requirement_import_preview_reports_missing_required_cells(client: TestClient):
    response = client.post(
        "/api/v1/requirements/import/preview",
        files={"file": _xlsx([["项目名称", "需求标题"], ["", ""]])},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["valid_count"] == 0
    assert data["error_count"] == 1
    assert data["errors"][0]["row_number"] == 2
    assert any("项目名称" in message for message in data["errors"][0]["messages"])
    assert any("需求标题" in message for message in data["errors"][0]["messages"])


def test_requirement_import_preview_detects_duplicate_title_in_project(client: TestClient):
    project_name = f"导入项目-{uuid4().hex[:8]}"
    project = client.post("/api/v1/projects", json={"name": project_name}).json()
    existing = client.post(
        "/api/v1/requirements",
        json={"project_id": project["id"], "title": "重复需求"},
    ).json()

    response = client.post(
        "/api/v1/requirements/import/preview",
        files={"file": _xlsx([["项目名称", "需求标题", "优先级"], [project_name, "重复需求", "3"]])},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["valid_count"] == 1
    assert data["duplicate_count"] == 1
    assert data["duplicates"][0]["row_number"] == 2
    assert data["duplicates"][0]["existing_requirement_id"] == existing["id"]


def test_requirement_import_preview_rejects_invalid_type(client: TestClient):
    project_name = f"导入类型项目-{uuid4().hex[:8]}"
    client.post("/api/v1/projects", json={"name": project_name}).json()

    response = client.post(
        "/api/v1/requirements/import/preview",
        files={"file": _xlsx([["项目名称", "需求标题", "类型"], [project_name, "新需求", "不存在类型"]])},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["valid_count"] == 0
    assert data["error_count"] == 1
    assert any("类型必须是" in message for message in data["errors"][0]["messages"])


def test_requirement_import_commit_creates_new_requirements(client: TestClient):
    project_name = f"导入新增项目-{uuid4().hex[:8]}"
    project = client.post("/api/v1/projects", json={"name": project_name}).json()

    response = client.post(
        "/api/v1/requirements/import/commit",
        data={"duplicate_strategy": "create_duplicate"},
        files={"file": _xlsx([["项目名称", "需求标题", "优先级"], [project_name, "新需求", "2"]])},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["created_count"] == 1
    assert data["updated_count"] == 0
    created = client.get("/api/v1/requirements").json()[0]
    assert created["project_id"] == project["id"]
    assert created["title"] == "新需求"
    assert created["priority"] == "2"


def test_project_scoped_requirement_import_creates_rows_in_current_project(client: TestClient):
    scoped_project = client.post("/api/v1/projects", json={"name": f"Scoped Import-{uuid4().hex[:8]}"}).json()
    other_project = client.post("/api/v1/projects", json={"name": f"Other Import-{uuid4().hex[:8]}"}).json()

    response = client.post(
        "/api/v1/requirements/import/commit",
        data={"duplicate_strategy": "create_duplicate", "project_id": str(scoped_project["id"])},
        files={
            "file": _xlsx(
                [
                    ["项目名称", "需求标题", "优先级"],
                    [other_project["name"], "当前项目导入需求 A", "2"],
                    ["", "当前项目导入需求 B", "4"],
                ]
            )
        },
    )

    assert response.status_code == 200
    assert response.json()["created_count"] == 2
    scoped_page = client.get(f"/api/v1/projects/{scoped_project['id']}/requirements?page=1&page_size=10").json()
    other_page = client.get(f"/api/v1/projects/{other_project['id']}/requirements?page=1&page_size=10").json()
    assert scoped_page["total"] == 2
    assert {item["title"] for item in scoped_page["items"]} == {"当前项目导入需求 A", "当前项目导入需求 B"}
    assert {item["project_id"] for item in scoped_page["items"]} == {scoped_project["id"]}
    assert other_page["total"] == 0


def test_requirement_import_commit_updates_duplicate_requirement(client: TestClient):
    project_name = f"导入更新项目-{uuid4().hex[:8]}"
    project = client.post("/api/v1/projects", json={"name": project_name}).json()
    existing = client.post(
        "/api/v1/requirements",
        json={"project_id": project["id"], "title": "重复需求", "priority": "3"},
    ).json()

    response = client.post(
        "/api/v1/requirements/import/commit",
        data={"duplicate_strategy": "update_existing"},
        files={"file": _xlsx([["项目名称", "需求标题", "优先级"], [project_name, "重复需求", "1"]])},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["created_count"] == 0
    assert data["updated_count"] == 1
    updated = client.get(f"/api/v1/requirements/{existing['id']}").json()
    assert updated["priority"] == "1"


def test_requirement_import_update_closes_active_iteration_history(client: TestClient):
    project_name = f"Import planned project-{uuid4().hex[:8]}"
    project = client.post("/api/v1/projects", json={"name": project_name}).json()
    iteration = client.post(
        "/api/v1/iterations",
        json={"project_ids": [project["id"]], "name": f"Import active history-{uuid4().hex[:8]}"},
    ).json()
    existing = client.post(
        "/api/v1/requirements",
        json={"project_id": project["id"], "iteration_id": iteration["id"], "title": "Imported planned requirement"},
    ).json()

    response = client.post(
        "/api/v1/requirements/import/commit",
        data={"duplicate_strategy": "update_existing"},
        files={"file": _xlsx([REQUIREMENT_IMPORT_COLUMNS, [project_name, "Imported planned requirement"]])},
    )

    assert response.status_code == 200
    assert client.get(f"/api/v1/requirements/{existing['id']}").json()["iteration_id"] is None
    db = SessionLocal()
    try:
        history = db.query(WorkItemIterationHistory).filter(
            WorkItemIterationHistory.object_type == "requirement",
            WorkItemIterationHistory.object_id == existing["id"],
        ).one()
        assert history.left_at is not None
        assert history.leave_reason == "import_updated"
        assert history.left_by is not None
    finally:
        db.close()


def test_requirement_import_update_rejects_terminal_iteration(client: TestClient):
    project_name = f"Import terminal project-{uuid4().hex[:8]}"
    project = client.post("/api/v1/projects", json={"name": project_name}).json()
    iteration = client.post(
        "/api/v1/iterations",
        json={"project_ids": [project["id"]], "name": f"Import terminal-{uuid4().hex[:8]}"},
    ).json()
    existing = client.post(
        "/api/v1/requirements",
        json={"project_id": project["id"], "iteration_id": iteration["id"], "title": "Terminal imported requirement"},
    ).json()
    db = SessionLocal()
    try:
        iteration_row = db.query(Iteration).filter(Iteration.id == iteration["id"]).one()
        terminal_state = db.query(WorkflowState).filter(
            WorkflowState.definition_id == iteration_row.workflow_definition_id,
            WorkflowState.category == "terminal",
        ).first()
        iteration_row.current_state_id = terminal_state.id
        db.commit()
    finally:
        db.close()

    response = client.post(
        "/api/v1/requirements/import/commit",
        data={"duplicate_strategy": "update_existing"},
        files={"file": _xlsx([REQUIREMENT_IMPORT_COLUMNS, [project_name, "Terminal imported requirement"]])},
    )

    assert response.status_code == 409
    assert response.json()["detail"]["code"] == "ITERATION_NOT_MUTABLE"
    assert client.get(f"/api/v1/requirements/{existing['id']}").json()["iteration_id"] == iteration["id"]


def test_requirement_import_commit_keeps_created_requirement_unassigned(client: TestClient):
    db = SessionLocal()
    try:
        owner = User(
            username=f"import_owner_{uuid4().hex[:8]}",
            full_name="导入默认负责人",
            password_hash=get_password_hash("User123456"),
            is_active=True,
        )
        db.add(owner)
        db.commit()
        owner_id = owner.id
    finally:
        db.close()
    config = client.post(
        "/api/v1/assignee-rule-configs",
        json={
            "name": f"导入负责人规则-{uuid4().hex[:8]}",
            "requirement_owner_roles": "tester",
            "creation_mode": "template",
            "template_source": {"source_type": "system", "source_id": "system-standard"},
        },
    ).json()
    assert client.post(f"/api/v1/assignee-rule-configs/{config['id']}/enable").status_code == 200
    project_name = f"导入负责人项目-{uuid4().hex[:8]}"
    project = client.post(
        "/api/v1/projects",
        json={"name": project_name, "assignee_rule_config_id": config["id"]},
    ).json()
    client.put(
        f"/api/v1/projects/{project['id']}/members",
        json=[{"user_id": owner_id, "project_role": "tester", "sort_order": 0}],
    )

    response = client.post(
        "/api/v1/requirements/import/commit",
        data={"duplicate_strategy": "create_duplicate"},
        files={"file": _xlsx([["项目名称", "需求标题", "优先级"], [project_name, "工作流负责人需求", "2"]])},
    )

    assert response.status_code == 200
    data = response.json()
    assert data["created_count"] == 1
    created = client.get("/api/v1/requirements").json()[0]
    assert created["title"] == "工作流负责人需求"
    assert created["owner_id"] is None


def test_requirement_import_update_preserves_existing_current_handler(client: TestClient):
    db = SessionLocal()
    try:
        old_owner = User(
            username=f"import_old_owner_{uuid4().hex[:8]}",
            full_name="导入旧负责人",
            password_hash=get_password_hash("User123456"),
            is_active=True,
        )
        new_owner = User(
            username=f"import_new_owner_{uuid4().hex[:8]}",
            full_name="导入新负责人",
            password_hash=get_password_hash("User123456"),
            is_active=True,
        )
        db.add_all([old_owner, new_owner])
        db.commit()
        old_owner_id = old_owner.id
        new_owner_id = new_owner.id
    finally:
        db.close()
    config = client.post(
        "/api/v1/assignee-rule-configs",
        json={
            "name": f"导入更新负责人规则-{uuid4().hex[:8]}",
            "requirement_owner_roles": "tester",
            "creation_mode": "template",
            "template_source": {"source_type": "system", "source_id": "system-standard"},
        },
    ).json()
    assert client.post(f"/api/v1/assignee-rule-configs/{config['id']}/enable").status_code == 200
    project_name = f"导入更新负责人项目-{uuid4().hex[:8]}"
    project = client.post(
        "/api/v1/projects",
        json={"name": project_name, "assignee_rule_config_id": config["id"]},
    ).json()
    client.put(
        f"/api/v1/projects/{project['id']}/members",
        json=[{"user_id": new_owner_id, "project_role": "tester", "sort_order": 0}],
    )
    existing = client.post(
        "/api/v1/requirements",
        json={"project_id": project["id"], "title": "重复工作流负责人需求", "owner_id": old_owner_id},
    ).json()

    response = client.post(
        "/api/v1/requirements/import/commit",
        data={"duplicate_strategy": "update_existing"},
        files={"file": _xlsx([["项目名称", "需求标题", "优先级"], [project_name, "重复工作流负责人需求", "1"]])},
    )

    assert response.status_code == 200
    updated = client.get(f"/api/v1/requirements/{existing['id']}").json()
    assert updated["owner_id"] == old_owner_id
