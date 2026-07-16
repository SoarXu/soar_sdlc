from dataclasses import dataclass
from io import BytesIO

from fastapi import HTTPException, status
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy.orm import Session

from app.models.project import Project
from app.models.requirement import Requirement
from app.models.user import User
from app.services.lifecycle_service import project_lifecycle_phase
from app.services.workflow_state_service import initial_workflow_values
from app.services.project_permission_service import ensure_work_item_create_permission


REQUIREMENT_IMPORT_COLUMNS = [
    "项目名称",
    "需求标题",
    "类型",
    "优先级",
    "提出人",
    "需求描述",
    "验收标准",
]
REQUIRED_COLUMNS = ["项目名称", "需求标题"]
REQUIREMENT_TYPE_VALUES = ["功能", "接口", "性能", "安全", "体验", "改进", "其他"]
PRIORITY_VALUES = {"1", "2", "3", "4", "5"}


@dataclass
class ParsedRequirementRow:
    row_number: int
    project_id: int
    project_name: str
    title: str
    requirement_type: str | None
    priority: str
    owner_id: int | None
    proposer_id: int | None
    review_status: str
    description: str | None
    acceptance_criteria: str | None


def build_requirement_import_template(db: Session, project_id: int | None = None) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "需求导入"
    sheet.append(REQUIREMENT_IMPORT_COLUMNS)
    project_name = _template_project_name(db, project_id)
    sheet.append([project_name, "示例需求", "功能", "3", "", "需求描述", "验收标准"])
    _add_list_validation(sheet, "C", REQUIREMENT_TYPE_VALUES)
    _add_list_validation(sheet, "D", sorted(PRIORITY_VALUES))
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = 18
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def ensure_excel_filename(filename: str) -> None:
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="仅支持 .xlsx 文件")


def preview_requirement_import(
    db: Session,
    file_bytes: bytes,
    project_id: int | None = None,
    actor: User | None = None,
) -> dict:
    parsed_rows, errors = _parse_requirement_rows(db, file_bytes, project_id)
    _ensure_rows_create_permission(db, parsed_rows, actor)
    duplicates = _duplicate_rows(db, parsed_rows)
    return {
        "valid_count": len(parsed_rows),
        "error_count": len(errors),
        "duplicate_count": len(duplicates),
        "errors": errors,
        "duplicates": duplicates,
    }


def commit_requirement_import(
    db: Session,
    file_bytes: bytes,
    duplicate_strategy: str,
    project_id: int | None = None,
    actor: User | None = None,
) -> dict:
    if duplicate_strategy not in {"update_existing", "create_duplicate"}:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="未知重复处理策略")
    parsed_rows, errors = _parse_requirement_rows(db, file_bytes, project_id)
    _ensure_rows_create_permission(db, parsed_rows, actor)
    if errors:
        return {"created_count": 0, "updated_count": 0, "error_count": len(errors), "errors": errors}

    created_count = 0
    updated_count = 0
    for row in parsed_rows:
        existing = _find_existing_requirement(db, row.project_id, row.title)
        if existing and duplicate_strategy == "update_existing":
            _apply_row_to_requirement(db, existing, row)
            updated_count += 1
            continue
        workflow_values = initial_workflow_values(db, "requirement", row.project_id)
        requirement = Requirement(
            project_id=row.project_id,
            source_project_id=None,
            iteration_id=None,
            title=row.title,
            requirement_type=row.requirement_type,
            priority=row.priority,
            owner_id=row.owner_id,
            proposer_id=row.proposer_id,
            **workflow_values,
            review_status=row.review_status,
            lifecycle_phase=project_lifecycle_phase(db, row.project_id),
            description=row.description,
            acceptance_criteria=row.acceptance_criteria,
        )
        db.add(requirement)
        created_count += 1
    db.commit()
    return {"created_count": created_count, "updated_count": updated_count, "error_count": 0, "errors": []}


def _parse_requirement_rows(
    db: Session,
    file_bytes: bytes,
    project_id: int | None = None,
) -> tuple[list[ParsedRequirementRow], list[dict]]:
    try:
        workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Excel 文件无法解析") from exc
    sheet = workbook.active
    headers = {str(cell.value).strip(): index for index, cell in enumerate(sheet[1]) if cell.value}
    required_columns = REQUIRED_COLUMNS if project_id is None else [REQUIRED_COLUMNS[1]]
    missing_columns = [column for column in required_columns if column not in headers]
    if missing_columns:
        return [], [{"row_number": 1, "messages": [f"缺少列：{column}" for column in missing_columns]}]

    scoped_project = _resolve_project_id(db, project_id) if project_id is not None else None
    rows = []
    errors = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = {header: _cell_text(row[index].value) for header, index in headers.items()}
        parsed, messages = _parse_requirement_row(db, row_number, values, scoped_project)
        if messages:
            errors.append({"row_number": row_number, "messages": messages})
        else:
            rows.append(parsed)
    return rows, errors


def _parse_requirement_row(
    db: Session,
    row_number: int,
    values: dict[str, str],
    scoped_project: Project | None = None,
) -> tuple[ParsedRequirementRow | None, list[str]]:
    messages = []
    project_name = values.get("项目名称", "")
    title = values.get("需求标题", "")
    if scoped_project is None and not project_name:
        messages.append("项目名称不能为空")
    if not title:
        messages.append("需求标题不能为空")

    project = scoped_project or (_resolve_project(db, project_name, "项目名称", messages) if project_name else None)
    proposer = _resolve_user(db, values.get("提出人", ""), "提出人", messages)
    requirement_type = values.get("类型") or None
    if requirement_type and requirement_type not in REQUIREMENT_TYPE_VALUES:
        messages.append(f"类型必须是 {'、'.join(REQUIREMENT_TYPE_VALUES)}")
    priority = values.get("优先级") or "3"
    if priority not in PRIORITY_VALUES:
        messages.append("优先级必须是 1、2、3、4、5")
    if messages:
        return None, messages
    return (
        ParsedRequirementRow(
            row_number=row_number,
            project_id=project.id,
            project_name=project.name,
            title=title,
            requirement_type=requirement_type,
            priority=priority,
            owner_id=None,
            proposer_id=proposer.id if proposer else None,
            review_status="not_required",
            description=values.get("需求描述") or None,
            acceptance_criteria=values.get("验收标准") or None,
        ),
        [],
    )


def _resolve_project(db: Session, name: str, field: str, messages: list[str]) -> Project | None:
    if not name:
        return None
    matches = db.query(Project).filter(Project.name == name, Project.deleted == 0).all()
    if not matches:
        messages.append(f"{field}不存在：{name}")
        return None
    if len(matches) > 1:
        messages.append(f"{field}不唯一：{name}")
        return None
    return matches[0]


def _resolve_project_id(db: Session, project_id: int) -> Project:
    project = db.query(Project).filter(Project.id == project_id, Project.deleted == 0).first()
    if not project:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="项目不存在")
    return project


def _resolve_user(db: Session, value: str, field: str, messages: list[str]) -> User | None:
    if not value:
        return None
    matches = (
        db.query(User)
        .filter(
            User.deleted == 0,
            User.is_active.is_(True),
            (User.full_name == value) | (User.username == value),
        )
        .all()
    )
    if not matches:
        messages.append(f"{field}不存在：{value}")
        return None
    if len(matches) > 1:
        messages.append(f"{field}不唯一：{value}")
        return None
    return matches[0]


def _duplicate_rows(db: Session, rows: list[ParsedRequirementRow]) -> list[dict]:
    duplicates = []
    for row in rows:
        existing = _find_existing_requirement(db, row.project_id, row.title)
        if existing:
            duplicates.append(
                {
                    "row_number": row.row_number,
                    "project_id": row.project_id,
                    "project_name": row.project_name,
                    "title": row.title,
                    "existing_requirement_id": existing.id,
                    "existing_requirement_title": existing.title,
                }
            )
    return duplicates


def _find_existing_requirement(db: Session, project_id: int, title: str) -> Requirement | None:
    return (
        db.query(Requirement)
        .filter(Requirement.project_id == project_id, Requirement.title == title, Requirement.deleted == 0)
        .first()
    )


def _apply_row_to_requirement(db: Session, requirement: Requirement, row: ParsedRequirementRow) -> None:
    requirement.source_project_id = None
    requirement.iteration_id = None
    requirement.requirement_type = row.requirement_type
    requirement.priority = row.priority
    if row.owner_id is not None:
        requirement.owner_id = row.owner_id
    requirement.proposer_id = row.proposer_id
    requirement.review_status = row.review_status
    requirement.description = row.description
    requirement.acceptance_criteria = row.acceptance_criteria


def _ensure_rows_create_permission(db: Session, rows: list[ParsedRequirementRow], actor: User | None) -> None:
    checked_project_ids: set[int] = set()
    for row in rows:
        if row.project_id in checked_project_ids:
            continue
        ensure_work_item_create_permission(db, row.project_id, actor)
        checked_project_ids.add(row.project_id)


def _cell_text(value) -> str:
    return str(value).strip() if value is not None else ""


def _template_project_name(db: Session, project_id: int | None) -> str:
    if not project_id:
        return "示例项目"
    project = db.query(Project).filter(Project.id == project_id, Project.deleted == 0).first()
    return project.name if project else "示例项目"


def _add_list_validation(sheet, column_letter: str, values: list[str]) -> None:
    validation = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
    sheet.add_data_validation(validation)
    validation.add(f"{column_letter}2:{column_letter}500")
